import csv
from openpyxl import load_workbook
import shutil

def cleanHousingData():
    """ This function creates two CSV files that includes data about Housing; Vacancy Rates and Rent Prices.
    The format of the data is as follows: Location, Bachelor, 1  Bedroom, 2 Bedroom, 3 Bedroom + """

    def columnAverages(filepath, sheetName, column1, column2, cellStartRow):

        """ This function creates a list for each column in the final CSV files based on average of the two data
        provided. """

        # reading excel files setup

        workbook = load_workbook(filename=filepath, data_only=True)

        spreadsheet = workbook[sheetName]

        # creates dictionary of Oct 18 cells

        cellsOct18 = {}

        cellRow = cellStartRow

        for i in range(72):
            cellsOct18["18cell" + str(i)] = spreadsheet[column1 + str(cellRow)].value
            cellRow += 1

        # creates dictionary of Oct 19 cells

        cellsOct19 = {}

        cellRow = cellStartRow

        for i in range(72):
            cellsOct19["19cell" + str(i)] = spreadsheet[column2 + str(cellRow)].value
            cellRow += 1

        # combine column as averages in one column called "averagesList"

        averagesList = {}

        n = 0

        for (i, e) in zip(cellsOct18.values(), cellsOct19.values()):

            # missing data is equated to zero or ignored

            if i == '**' and e == '**':
                averagesList["averagecell" + str(n)] = 0

            elif i == '**':
                averagesList["averagecell" + str(n)] = e

            elif e == '**':
                averagesList["averagecell" + str(n)] = i

            else:
                averagesList["averagecell" + str(n)] = round((i + e)/2, 2)

            n += 1

        return averagesList.values()

    def locationDic(filepath, sheetName, cellStartRow):

        """ This function creates a dictionary with the locations of the houses. """

        #reading excel files setup

        workbook = load_workbook(filename=filepath, data_only=True)

        spreadsheet = workbook[sheetName]

        # creates a dictionary for locations

        locationDic = {}

        for i in range(72):
            locationDic["location" + str(i)] = spreadsheet["A" + str(cellStartRow)].value
            cellStartRow += 1

        # returns list of values of dictionary

        return locationDic.values()

    # original housing data file path

    housingDataPath = "workingData/HousingData/OntarioHousingData.xlsx"

    def convertHousingVR():
        """ This function applies above functions on the orignal housing data, then, creates an CSV file with the
        housing vacancy rates data. """

        housingSheet1 = "Table 3.1.1"

        # apply locationDic function on Vacancy Rates data

        locationVR = locationDic(housingDataPath, housingSheet1, 8)

        # apply columnAverages function on Vacancy Rates data

        sheetVRBachelor = columnAverages(housingDataPath, housingSheet1, "B", "D", 8)
        sheetVROneB = columnAverages(housingDataPath, housingSheet1, "G", "I", 8)
        sheetVRTwoB = columnAverages(housingDataPath, housingSheet1, "L", "N", 8)
        sheetVRThreeB = columnAverages(housingDataPath, housingSheet1, "Q", "S", 8)

        # create a new CSV file with the housing vacancy rates data

        with open('housingVacancyRates.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Location", "Bachelor", "1 Bedroom", "2 Bedroom", "3 Bedroom +"])

            for (loc, Bach, OneB, TwoB, ThreeB) in zip(locationVR, sheetVRBachelor, sheetVROneB, sheetVRTwoB,
                                                       sheetVRThreeB):
                writer.writerow([loc, Bach, OneB, TwoB, ThreeB])

        # move file from current directory to specified one for data

        shutil.move("housingVacancyRates.csv", "workingData/HousingData/housingVacancyRates.csv")


    def convertHousingRP():
        """ This function applies above functions on the orignal housing data, then, creates an CSV file with the
                housing rent prices data. """

        housingSheet2 = "Table 3.1.2"

        # apply locationDic function on Vacancy Rates data

        locationRP = locationDic(housingDataPath, housingSheet2, 7)

        # apply columnAverages function on Rent Prices data

        sheetRPBachelor = columnAverages(housingDataPath, housingSheet2, "B", "D", 7)
        sheetRPOneB = columnAverages(housingDataPath, housingSheet2, "F", "H", 7)
        sheetRPTwoB = columnAverages(housingDataPath, housingSheet2, "J", "L", 7)
        sheetRPThreeB = columnAverages(housingDataPath, housingSheet2, "N", "P", 7)

        # create a new CSV file with the housing rent prices data

        with open('housingRentPrices.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Location", "Bachelor", "1 Bedroom", "2 Bedroom", "3 Bedroom +"])

            for (loc, Bach, OneB, TwoB, ThreeB) in zip(locationRP, sheetRPBachelor, sheetRPOneB, sheetRPTwoB, sheetRPThreeB):
                writer.writerow([loc, Bach, OneB, TwoB, ThreeB])

        # move file from current directory to specified one for data

        shutil.move("housingRentPrices.csv", "workingData/HousingData/housingRentPrices.csv")

    convertHousingVR()

    convertHousingRP()

def cleanEducationData():
    """ This function creates an CSV file that includes data about Education.
        The format of the data is as follows: Location, schoolName, schoolType, schoolLevel, gradeRange """

    # reading excel files setup

    educationWorkbook = load_workbook(filename="workingData/SchoolsData/publicly_funded_schools_xlsx_september_2020_en.xlsx", data_only=True)

    educationSpreadsheet = educationWorkbook["EN"]

    # initialize/define variables for column lists

    schoolNames = {}
    schoolLocations = {}
    schoolType = {}
    schoolLevel = {}
    gradeRange = {}

    schoolRow = 2
    schoolListNumber = 0

    for i in range(5850):

        if educationSpreadsheet["K" + str(schoolRow)].value == "Not applicable":

            schoolDictKey = "school" + str(schoolListNumber)

            # create lists for columns

            schoolNames[schoolDictKey] = educationSpreadsheet["G" + str(schoolRow)].value
            schoolLocations[schoolDictKey] = educationSpreadsheet["O" + str(schoolRow)].value
            schoolType[schoolDictKey] = educationSpreadsheet["J" + str(schoolRow)].value
            schoolLevel[schoolDictKey] = educationSpreadsheet["H" + str(schoolRow)].value
            gradeRange[schoolDictKey] = educationSpreadsheet["T" + str(schoolRow)].value


        schoolRow += 1

        schoolListNumber += 1

    # create a file with the education data

    with open('educationData.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["schoolLocation", "schoolName", "schoolType", "schoolLevel", "gradeRange"])

        for (loc, name, ty, level, grades) in zip(schoolLocations.values(), schoolNames.values(), schoolType.values(), schoolLevel.values(), gradeRange.values()):
            writer.writerow([loc, name, ty, level, grades])

    # move file from current directory to specified one for data

    shutil.move("educationData.csv", "workingData/SchoolsData/educationData.csv")

def cleanHealthcareData():
    """ This function creates an CSV file that includes data about Healthcare.
            The format of the data is as follows: Location, healthcareName, serviceType """

    output_healthcareDataFile = "HealthcareData.csv"

    healthcareName = {}
    serviceType = {}
    healthcareLocation = {}

    healthcareDictKey = 0


    with open("workingData/HealthcareData/Ministry_of_Health_Service_Provider_Locations.csv") as csv_file:

        reader = csv.reader(csv_file, delimiter=',')

        for row in reader:

            if healthcareDictKey >= 1:

                healthcareName["heathcareName" + str(healthcareDictKey)] = row[6]
                serviceType["serviceType" + str(healthcareDictKey)] = row[4]
                healthcareLocation["location" + str(healthcareDictKey)] = row[13]

            healthcareDictKey += 1

        # place data in lists

        with open(output_healthcareDataFile, "w", newline='') as result:

            writer = csv.writer(result)

            writer.writerow(["Location", "HealthcareName", "ServiceType"])

            for(loc, name, ty) in zip(healthcareLocation.values(), healthcareName.values(), serviceType.values()):
                writer.writerow([loc, name, ty])

        # move file to specified folder for data

        shutil.move("HealthcareData.csv", "workingData/HealthcareData/HealthcareData.csv")


cleanHousingData()
cleanEducationData()
cleanHealthcareData()